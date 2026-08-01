"""Typed, bounded descriptors for user-attached data files.

The agent must learn a file's SHAPE, never its CONTENT: the generated app loads the full data at
runtime from a served URL, so nothing here is ever the app's data source. The previous approach —
inlining a utf-8-decoded 8KB head labelled "SCHEMA SAMPLE" — only worked by accident for CSV. A PDF
or PNG decodes to mojibake (U+FFFD everywhere), which burns prompt context and actively misleads
the model about the file's type. So: sniff the real type, then emit a DERIVED description.

Type detection is by magic bytes first and extension only as a fallback — these files are
user-uploaded and the names are untrustworthy (an .xlsx that is really a CSV is common).

`describe()` never raises. It is called while assembling a prompt, where a bad attachment must
degrade to a line of honest text rather than fail the turn — hence the blanket excepts.

openpyxl / pyarrow / pypdf are NOT dependencies of this backend. Each is imported inside the branch
that needs it, and an ImportError degrades to the lesser descriptor we can still produce from the
file's bytes (kind + size), never to a crash.
"""
from __future__ import annotations

import csv
import io
import json
import os
import re
import struct
from pathlib import Path

# Enough to cover every magic number we check plus a representative CSV/text head. Read once.
_HEAD_BYTES = 8192
# csv sample rows used for type inference. More rows barely improve the guess and cost a re-scan.
_TYPE_SAMPLE_ROWS = 50
# Depth/width bounds on inferred JSON schemas. A 50MB document's schema is a few hundred chars.
_JSON_MAX_DEPTH = 4
_JSON_MAX_KEYS = 60
# json.load holds the whole document in memory, so refuse outright above this and say so.
_JSON_PARSE_LIMIT = 32 * 1024 * 1024
# Extracting text is the expensive part of a PDF; per-page counts past this add nothing.
_PDF_MAX_PAGES_SCANNED = 20
_SUMMARY_MAX = 90

_DATE_RE = re.compile(r"^\d{4}-\d{2}-\d{2}([ T]\d{2}:\d{2}(:\d{2})?)?$|^\d{1,2}/\d{1,2}/\d{2,4}$")


def describe(path: str, *, max_detail_chars: int = 1200) -> dict:
    """Describe the file at `path` as {kind, summary, detail, size}.

    `summary` is a single line <= 90 chars — it is rendered into an always-present index of every
    attachment, so it must be cheap enough to show unconditionally. `detail` is a multi-line block
    hard-capped at `max_detail_chars`, inlined only when the user explicitly mentions that file.
    """
    try:
        size = os.path.getsize(path)
    except OSError as e:
        return _unavailable(f"cannot stat file — {e.strerror or type(e).__name__}")
    try:
        with open(path, "rb") as f:
            head = f.read(_HEAD_BYTES)
    except OSError as e:
        return _unavailable(f"cannot read file — {e.strerror or type(e).__name__}", size)

    try:
        kind, hint = _sniff(path, head)
    except Exception as e:  # a malformed file must never take down the prompt assembly
        return _unavailable(f"could not be identified — {type(e).__name__}", size)
    try:
        summary, detail = _HANDLERS[kind](path, head, hint, size)
    except Exception as e:
        # Magic bytes already identified the type, and that identification is reliable — a parse
        # failure must degrade the DESCRIPTION, not the identification. A truncated PDF is still a
        # PDF, and telling the agent that beats a useless "unavailable".
        summary = f"{kind} — {_human(size)}, could not be parsed"
        detail = (f"Identified as {kind} by its magic bytes, but parsing failed "
                  f"({type(e).__name__}) — the file may be truncated or corrupt. "
                  f"Content was NOT previewed.")

    return {"kind": kind, "summary": _one_line(summary), "detail": _cap(detail, max_detail_chars),
            "size": size}


# ---------------------------------------------------------------- detection

def _sniff(path: str, head: bytes) -> tuple[str, object]:
    """Return (kind, hint) where hint is whatever the handler needs that we already computed here.

    Magic bytes win. Extension is consulted only to disambiguate containers that share a magic
    number (zip is xlsx, jar, or plain zip) and to pick a delimiter for text that is already known
    to be tabular.
    """
    ext = os.path.splitext(path)[1].lower()

    if head.startswith(b"%PDF"):
        return "pdf", None
    if head.startswith(b"PAR1"):
        return "parquet", None
    if head[:4] in (b"\x89PNG", b"GIF8") or head[:3] == b"\xff\xd8\xff" or (
            head[:4] == b"RIFF" and head[8:12] == b"WEBP"):
        return "image", None
    if head.startswith(b"PK\x03\x04"):
        return ("excel", None) if ext in (".xlsx", ".xlsm") else ("binary", "zip archive")

    text = _decode_head(head)
    if text is None:
        return "binary", None

    stripped = text.lstrip()
    if stripped[:1] in ("{", "["):
        return "json", None
    if _looks_tabular(text, ext):
        return "tabular", "\t" if ext == ".tsv" else _delimiter(text)
    return "text", text


def _decode_head(head: bytes) -> str | None:
    """Strict-decode the head, tolerating a multi-byte character sliced by the read boundary.

    Strict is the point: `errors="replace"` is exactly the bug this module exists to fix, since it
    turns any binary file into plausible-looking text full of U+FFFD.
    """
    for trim in range(4):
        try:
            return head[:len(head) - trim].decode("utf-8")
        except UnicodeDecodeError:
            continue
    return None


def _looks_tabular(text: str, ext: str) -> bool:
    if ext in (".csv", ".tsv"):
        return True
    lines = [ln for ln in text.splitlines()[:5] if ln.strip()]
    if len(lines) < 2:
        return False
    d = _delimiter(text)
    counts = {ln.count(d) for ln in lines}
    return len(counts) == 1 and counts.pop() >= 1


def _delimiter(text: str) -> str:
    try:
        return csv.Sniffer().sniff(text[:4096], delimiters=",\t;|").delimiter
    except csv.Error:
        return ","


# ---------------------------------------------------------------- handlers

def _describe_tabular(path: str, head: bytes, delim: str, size: int) -> tuple[str, str]:
    """Columns with inferred types, a row count, and 2-3 sample rows.

    Deliberately derived rather than raw head lines: a 200-column file's head lines are enormous,
    and no number of them tells the agent how many rows exist.
    """
    with open(path, encoding="utf-8", errors="ignore", newline="") as f:
        reader = csv.reader(f, delimiter=delim)
        try:
            header = next(reader)
        except StopIteration:
            return "Empty delimited file — no header row", ""
        sample = []
        for row in reader:
            sample.append(row)
            if len(sample) >= _TYPE_SAMPLE_ROWS:
                break
        rows = len(sample) + sum(1 for _ in reader)   # streams; never materializes the file

    types = [_column_type([r[i] for r in sample if i < len(r)]) for i in range(len(header))]
    label = "TSV" if delim == "\t" else "CSV"
    summary = f"{label} — {len(header)} columns, {rows:,} rows"

    lines = [f"{len(header)} columns, {rows:,} data rows, delimiter {delim!r}.", "Columns:"]
    lines += [f"  {name or f'(unnamed {i})'}: {t}" for i, (name, t) in enumerate(zip(header, types))]
    if sample:
        lines.append("Sample rows:")
        lines += [f"  {delim.join(r)}" for r in sample[:3]]
    return summary, "\n".join(lines)


def _column_type(values: list[str]) -> str:
    vals = [v.strip() for v in values if v and v.strip()]
    if not vals:
        return "string"
    for name, test in (("bool", _is_bool), ("int", _is_int), ("float", _is_float),
                       ("date", _DATE_RE.match)):
        if all(test(v) for v in vals):
            return name
    return "string"


def _is_bool(v: str) -> bool:
    return v.lower() in ("true", "false")


def _is_int(v: str) -> bool:
    return v.lstrip("+-").isdigit()


def _is_float(v: str) -> bool:
    try:
        float(v)
        return True
    except ValueError:
        return False


def _describe_json(path: str, head: bytes, hint, size: int) -> tuple[str, str]:
    """An inferred schema (key paths + types), never values.

    Values are the whole risk here: a JSON attachment can be 50MB of records, and any of them
    pasted into the prompt invites the model to hardcode them as the app's data.
    """
    if size > _JSON_PARSE_LIMIT:
        return (f"JSON — {_human(size)}, too large to parse for a schema",
                ("Schema not inferred: the document exceeds the in-memory parse limit. "
                 "The app should stream it at runtime from its served URL."))
    with open(path, encoding="utf-8", errors="strict") as f:
        text = f.read()

    try:
        doc = json.loads(text)
    except json.JSONDecodeError:
        records = _parse_ndjson(text)
        if not records:
            return "JSON — malformed, schema not inferred", "The file is not valid JSON or NDJSON."
        paths: list[str] = []
        for r in records:
            _walk(r, "", 1, paths)
        n = sum(1 for ln in text.splitlines() if ln.strip())
        return (f"NDJSON — {n:,} records, {len(paths)} fields",
                f"Newline-delimited JSON, {n:,} records. Schema inferred from the first "
                f"{len(records)} records:\n" + "\n".join(_dedupe(paths)))

    paths = []
    _walk(doc, "", 1, paths)
    if isinstance(doc, list):
        summary = f"JSON — array of {len(doc):,} items, {len(paths)} fields"
    elif isinstance(doc, dict):
        summary = f"JSON — object, {len(doc)} top-level keys"
    else:
        summary = f"JSON — single {_json_type(doc)} value"
    return summary, "Inferred schema (key paths and types, values omitted):\n" + "\n".join(_dedupe(paths))


def _parse_ndjson(text: str, limit: int = 20) -> list:
    out = []
    for line in text.splitlines():
        if not line.strip():
            continue
        try:
            out.append(json.loads(line))
        except json.JSONDecodeError:
            return []
        if len(out) >= limit:
            break
    return out


def _walk(node, prefix: str, depth: int, out: list[str]) -> None:
    if len(out) >= _JSON_MAX_KEYS:
        return
    if isinstance(node, dict):
        if depth > _JSON_MAX_DEPTH:
            out.append(f"{prefix}: object (nested deeper, not expanded)")
            return
        for k, v in node.items():
            path = f"{prefix}.{k}" if prefix else k
            if isinstance(v, (dict, list)):
                _walk(v, path, depth + 1, out)
            else:
                out.append(f"{path}: {_json_type(v)}")
            if len(out) >= _JSON_MAX_KEYS:
                return
    elif isinstance(node, list):
        label = prefix + "[]" if prefix else "[]"
        if not node:
            out.append(f"{label}: empty array")
        elif isinstance(node[0], (dict, list)) and depth <= _JSON_MAX_DEPTH:
            _walk(node[0], label, depth + 1, out)
        else:
            out.append(f"{label}: array of {_json_type(node[0])}")
    else:
        out.append(f"{prefix or '(root)'}: {_json_type(node)}")


def _json_type(v) -> str:
    if v is None:
        return "null"
    return {bool: "bool", int: "int", float: "float", str: "string"}.get(type(v), type(v).__name__)


def _dedupe(paths: list[str]) -> list[str]:
    seen, out = set(), []
    for p in paths:
        if p not in seen:
            seen.add(p)
            out.append(p)
    return out


def _describe_excel(path: str, head: bytes, hint, size: int) -> tuple[str, str]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return (f"Excel workbook — {_human(size)}, sheet details unavailable",
                ("openpyxl is not installed, so sheet names and dimensions could not be read. "
                 "The file is a valid .xlsx workbook."))

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        names = wb.sheetnames
        lines = []
        for name in names:
            ws = wb[name]
            dims = f"{ws.max_row or '?'} rows x {ws.max_column or '?'} columns"
            header = next(ws.iter_rows(max_row=1, values_only=True), ()) or ()
            cells = ", ".join(str(c) for c in header if c is not None)
            lines.append(f"  {name}: {dims}" + (f"\n    header: {cells}" if cells else ""))
    finally:
        wb.close()
    return f"Excel — {len(names)} sheets", "Sheets:\n" + "\n".join(lines)


def _describe_parquet(path: str, head: bytes, hint, size: int) -> tuple[str, str]:
    try:
        import pyarrow.parquet as pq
    except ImportError:
        return (f"Parquet file — {_human(size)}, schema unavailable",
                ("pyarrow is not installed, so the column schema could not be read from the footer. "
                 "The file is a valid Parquet file."))

    pf = pq.ParquetFile(path)
    schema = pf.schema_arrow
    rows = pf.metadata.num_rows if pf.metadata else 0
    cols = "\n".join(f"  {f.name}: {f.type}" for f in schema)
    return (f"Parquet — {len(schema.names)} columns, {rows:,} rows",
            f"{len(schema.names)} columns, {rows:,} rows.\nColumns:\n{cols}")


def _describe_pdf(path: str, head: bytes, hint, size: int) -> tuple[str, str]:
    """Page count, outline, a short first-page snippet, and per-page character counts.

    The per-page counts are the load-bearing part: all-zero counts mean a scanned, image-only PDF,
    so the agent knows up front that text extraction will yield nothing.
    """
    try:
        from pypdf import PdfReader
    except ImportError:
        return (f"PDF document — {_human(size)}, page details unavailable",
                ("pypdf is not installed, so page count and text extraction were skipped. "
                 "The file is a valid PDF."))

    reader = PdfReader(path)
    pages = reader.pages
    texts = []
    for p in pages[:_PDF_MAX_PAGES_SCANNED]:
        try:
            texts.append(p.extract_text() or "")
        except Exception:
            texts.append("")

    lines = [f"{len(pages)} pages."]
    titles = _outline_titles(reader)
    if titles:
        lines.append("Outline: " + "; ".join(titles[:10]))
    counts = ", ".join(f"p{i + 1}={len(t)}" for i, t in enumerate(texts))
    lines.append(f"Extractable characters per page (first {len(texts)}): {counts}")
    if not any(texts):
        lines.append("No extractable text — this is a scanned/image-only PDF; text extraction "
                     "will not work.")
    else:
        snippet = _one_line(next(t for t in texts if t.strip()))[:200]
        lines.append(f"First page snippet: {snippet}")
    return f"PDF — {len(pages)} pages", "\n".join(lines)


def _outline_titles(reader) -> list[str]:
    out: list[str] = []

    def walk(items):
        for it in items:
            if isinstance(it, list):
                walk(it)
            else:
                title = getattr(it, "title", None)
                if title:
                    out.append(str(title))

    try:
        walk(reader.outline)
    except Exception:
        return []
    return out


def _describe_image(path: str, head: bytes, hint, size: int) -> tuple[str, str]:
    """Format and pixel dimensions parsed straight from the header — no decode, no pixel data.

    Pillow is not a dependency and is not worth one here: PNG/GIF/JPEG/WEBP all carry their
    dimensions in a fixed header field, and an unparseable header degrades to format-only.
    """
    fmt, dims = _image_header(head)
    wh = f"{dims[0]}x{dims[1]}" if dims else "dimensions unknown"
    detail = (f"{fmt} image, {wh}, {_human(size)}.\n"
              "Pixel data is not previewed. The app should reference the image by its served URL.")
    return f"{fmt} image — {wh}", detail


_IMAGE_MIMES = {"PNG": "image/png", "JPEG": "image/jpeg", "GIF": "image/gif", "WEBP": "image/webp"}


def image_mime(path: str) -> str | None:
    """MIME type of an image attachment, or None if it isn't a recognized image.

    Images are the one attachment type whose pixels ARE the shape, so Sage inlines them into the
    agent's prompt as `data:<mime>;base64,...`. Verified against OpenCode 1.18.4: a data: URI
    reaches the model intact, while a file-path URI makes OpenCode emit malformed base64 ("media
    must contain valid base64"). Sniffed from the header — an upload's filename is user-supplied
    and untrustworthy.
    """
    try:
        with open(path, "rb") as f:
            head = f.read(_HEAD_BYTES)
    except OSError:
        return None
    return _IMAGE_MIMES.get(_image_header(head)[0])


# Longest edge we shrink an oversized image to. Vision models downsample anyway — Anthropic's own
# guidance puts the useful ceiling around here — so a larger image costs prompt budget without
# adding signal. Shrinking beats refusing: a phone photo or a hi-DPI screenshot is exactly the kind
# of thing a user attaches, and the alternative is an agent that cannot see it at all.
_MAX_IMAGE_EDGE = 1568


def _decodes(data: bytes) -> bool:
    """Whether the pixels actually decode. Without Pillow we can't tell, so assume they do rather
    than refuse every image — the failure then surfaces at the provider instead, which is worse but
    still honest, whereas refusing outright would break images that are perfectly fine."""
    try:
        from PIL import Image
    except ImportError:
        return True
    try:
        with Image.open(io.BytesIO(data)) as im:
            im.verify()
        return True
    except Exception:
        return False


def fit_image(path: str, max_bytes: int) -> tuple[bytes, str] | None:
    """(bytes, mime) for an image small enough to inline, shrinking it if needed. None if it can't.

    Returns the file untouched when it already fits. Otherwise re-encodes: PNG first (lossless,
    right for the screenshots and diagrams people actually attach), falling back to JPEG and then to
    progressively smaller edges when PNG can't get under budget. None when Pillow is missing or the
    file won't decode — callers must then tell the agent it cannot see the image rather than let it
    guess.
    """
    try:
        data = Path(path).read_bytes()
    except OSError:
        return None
    mime = _IMAGE_MIMES.get(_image_header(data[:_HEAD_BYTES])[0])
    if mime is None:
        return None
    if len(data) <= max_bytes:
        # Still verify: magic bytes prove the type, not that the pixels decode. A small corrupt
        # image would otherwise sail through untouched and fail at the provider, while the Data
        # panel had already told the user the agent could see it.
        return (data, mime) if _decodes(data) else None
    try:
        from PIL import Image
    except ImportError:
        return None
    try:
        edge = _MAX_IMAGE_EDGE
        for _ in range(4):
            with Image.open(io.BytesIO(data)) as im:
                im = im.convert("RGB")
                im.thumbnail((edge, edge))
                for fmt, out_mime, kw in (("PNG", "image/png", {"optimize": True}),
                                          ("JPEG", "image/jpeg", {"quality": 85, "optimize": True})):
                    buf = io.BytesIO()
                    im.save(buf, fmt, **kw)
                    if buf.tell() <= max_bytes:
                        return buf.getvalue(), out_mime
            edge //= 2
    except Exception:  # a decodable header is no guarantee the pixels decode
        return None
    return None


def _image_header(head: bytes) -> tuple[str, tuple[int, int] | None]:
    if head[:4] == b"\x89PNG":
        if head[12:16] == b"IHDR":
            return "PNG", struct.unpack(">II", head[16:24])
        return "PNG", None
    if head[:4] == b"GIF8":
        return "GIF", struct.unpack("<HH", head[6:10])
    if head[:3] == b"\xff\xd8\xff":
        return "JPEG", _jpeg_dims(head)
    if head[:4] == b"RIFF":
        return "WEBP", _webp_dims(head)
    return "image", None


def _jpeg_dims(head: bytes) -> tuple[int, int] | None:
    """Walk the JPEG marker chain to the first SOFn frame header, which carries height/width."""
    i = 2
    while i + 9 < len(head):
        if head[i] != 0xFF:
            i += 1
            continue
        marker, length = head[i + 1], struct.unpack(">H", head[i + 2:i + 4])[0]
        if 0xC0 <= marker <= 0xCF and marker not in (0xC4, 0xC8, 0xCC):
            h, w = struct.unpack(">HH", head[i + 5:i + 9])
            return w, h
        i += 2 + length
    return None


def _webp_dims(head: bytes) -> tuple[int, int] | None:
    tag = head[12:16]
    if tag == b"VP8X":
        w = int.from_bytes(head[24:27], "little") + 1
        h = int.from_bytes(head[27:30], "little") + 1
        return w, h
    if tag == b"VP8 ":
        w, h = struct.unpack("<HH", head[26:30])
        return w & 0x3FFF, h & 0x3FFF
    if tag == b"VP8L":
        bits = int.from_bytes(head[21:25], "little")
        return (bits & 0x3FFF) + 1, ((bits >> 14) & 0x3FFF) + 1
    return None


def _describe_text(path: str, head: bytes, text: str, size: int) -> tuple[str, str]:
    """The one kind where raw content is the right answer — bounded, and already known decodable."""
    lines = text.splitlines()
    return f"Text — {_human(size)}", "First lines:\n" + "\n".join(lines[:40])


def _describe_binary(path: str, head: bytes, hint: str | None, size: int) -> tuple[str, str]:
    what = hint or "unrecognized binary format"
    return (f"Binary — {_human(size)}, {what}",
            (f"{os.path.basename(path)}: {_human(size)}, {what}.\n"
             "Content was NOT previewed — the bytes are not text and decoding them would produce "
             "garbage. Treat this file as an opaque blob served by URL."))


_HANDLERS = {
    "tabular": _describe_tabular,
    "json": _describe_json,
    "excel": _describe_excel,
    "parquet": _describe_parquet,
    "pdf": _describe_pdf,
    "image": _describe_image,
    "text": _describe_text,
    "binary": _describe_binary,
}


# ---------------------------------------------------------------- shaping

def _unavailable(reason: str, size: int = 0) -> dict:
    return {"kind": "unavailable", "summary": _one_line(f"Unavailable — {reason}"), "detail": "",
            "size": size}


def _one_line(s: str) -> str:
    s = " ".join(s.split())
    return s if len(s) <= _SUMMARY_MAX else s[:_SUMMARY_MAX - 1] + "…"


def _cap(detail: str, limit: int) -> str:
    note = "\n… (truncated)"
    if len(detail) <= limit:
        return detail
    if limit <= len(note):
        return detail[:limit]
    return detail[:limit - len(note)] + note


def _human(n: int) -> str:
    for unit in ("B", "KB", "MB", "GB"):
        if n < 1024 or unit == "GB":
            return f"{n:.0f} {unit}" if unit == "B" else f"{n:.1f} {unit}"
        n /= 1024.0
    return f"{n} B"
